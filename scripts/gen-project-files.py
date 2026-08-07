# Generates the Seduh Coffee final-project artifacts in public/project/.
#
# The program runs 12 sessions and learners join mid-way, so every session page
# offers a "ready to continue" file: the project as it should look at the START
# of that session. The data pipeline only branches once (Session 2 cleaning), so
# 12 sessions need 5 derived artifacts, not 12.
#
#   seduh-coffee-final-project.xlsx   raw + messy      (patched in place here)
#   seduh-coffee-cleaned.xlsx         after Session 2
#   seduh-coffee-data-pack.zip        CSVs + schema.sql + seduh.db
#   seduh-coffee-analysis.xlsx        cleaned + the Q1-Q4 / Q7 answers
#   seduh-coffee-rfm-segments.csv     after Session 10
#   seduh-coffee-deck-outline.md      slide skeleton for Session 11
#
# The three workbooks that carry narrative tabs also get an -id sibling
# (…-id.xlsx) with only the Project Brief and Data Dictionary tabs translated to
# Bahasa Indonesia; every data tab stays English. Served when the site is set to
# Indonesian — see CONTINUATION_FILES in src/data/finalProject.ts.
#
# Dev-only: run by hand, commit the output. Nothing in the build invokes it.
#   pip install pandas openpyxl
#   python scripts/gen-project-files.py
#
# Re-running is safe — every write is an absolute value, never an append.

import shutil
import sqlite3
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / 'public' / 'project'
RAW = OUT / 'seduh-coffee-final-project.xlsx'

DATA_SHEETS = ['Orders', 'Customers', 'Products', 'Marketing_Spend']

# Sessions 7 (Power BI), 8 (Python) and 10 (RFM) are optional: the portfolio
# piece stands without them. Kept in sync with OPTIONAL_STEPS in
# src/data/finalProject.ts.
BULLET = '•'
EM = '—'

# RFM is anchored to the day after the data ends, per the brief.
SNAPSHOT = pd.Timestamp('2026-01-01')


# ── Session 2: the cleaning rules ────────────────────────────────────────────

# Every messy spelling in Orders.channel folds onto one of these four. Matching
# happens on the trimmed, whitespace-collapsed, lowercased value.
CHANNEL_MAP = {
    'webstore': 'Webstore',
    'tokopedia': 'Tokopedia',
    'tokopdia': 'Tokopedia',   # typo in the source
    'shopee': 'Shopee',
    'tiktok shop': 'TikTok Shop',
}

GENDER_MAP = {
    'female': 'Female', 'f': 'Female', 'perempuan': 'Female',
    'male': 'Male', 'm': 'Male', 'laki-laki': 'Male',
}

CATEGORY_MAP = {'Ready-to-Drink': 'Ready to Drink'}


def _trim(series):
    """Strip and collapse internal runs of whitespace, preserving nulls."""
    return series.where(series.isna(), series.astype('string').str.strip().str.replace(r'\s+', ' ', regex=True))


def clean(raw):
    """Apply the Session 2 rules. Returns (cleaned tables, cleaning log rows)."""
    log = []
    orders = raw['Orders'].copy()
    customers = raw['Customers'].copy()
    products = raw['Products'].copy()
    spend = raw['Marketing_Spend'].copy()

    def note(table, column, issue, rule, rows):
        log.append({
            'Step': len(log) + 1, 'Table': table, 'Column': column,
            'Issue': issue, 'Rule applied': rule, 'Rows affected': rows,
        })

    # --- Orders ---------------------------------------------------------
    # Text normalisation comes BEFORE de-duplication on purpose: one duplicate
    # pair differs only by the 'Tokopdia' typo, so normalising first turns it
    # into an exact duplicate and a single rule removes all 14.
    untrimmed = (orders['channel'] != orders['channel'].str.strip()).sum()
    variants = orders['channel'].nunique()
    orders['channel'] = _trim(orders['channel']).str.lower().map(CHANNEL_MAP)
    if orders['channel'].isna().any():
        raise SystemExit('Orders.channel has a spelling CHANNEL_MAP does not cover')
    note('Orders', 'channel',
         f'{variants} spellings of 4 real channels ({untrimmed} with stray spaces)',
         'TRIM, collapse spaces, lowercase, map to Webstore / Tokopedia / Shopee / TikTok Shop', untrimmed)

    for col in ['payment_method', 'city', 'province', 'order_status']:
        orders[col] = _trim(orders[col])

    before = len(orders)
    orders = orders.drop_duplicates().reset_index(drop=True)
    note('Orders', 'all columns', 'Order lines recorded twice',
         'Drop exact duplicate rows (after channel normalisation)', before - len(orders))

    if orders['order_id'].duplicated().any():
        raise SystemExit('Orders still has conflicting duplicate order_id values')

    before = len(orders)
    orders = orders[orders['quantity'] > 0].reset_index(drop=True)
    note('Orders', 'quantity', 'Zero or negative quantity is not a real sale',
         'Drop rows where quantity <= 0', before - len(orders))

    note('Orders', 'rating', f'{int(orders["rating"].isna().sum())} blank ratings',
         'LEFT AS BLANK — an unrated order is not a defect. Never imputed; '
         'exclude blanks when averaging rating.', 0)

    # --- Customers ------------------------------------------------------
    variants = customers['gender'].nunique()
    before_gender = customers['gender'].copy()
    customers['gender'] = _trim(customers['gender']).str.lower().map(GENDER_MAP)
    if customers['gender'].isna().any():
        raise SystemExit('Customers.gender has a value GENDER_MAP does not cover')
    note('Customers', 'gender', f'{variants} spellings across 2 values (EN and ID mixed)',
         'Lowercase and map to Female / Male',
         int((before_gender != customers['gender']).fillna(False).sum()))

    untrimmed = (customers['city'] != customers['city'].str.strip()).fillna(False).sum()
    customers['city'] = _trim(customers['city'])
    note('Customers', 'city', 'Trailing spaces split one city into two values',
         'TRIM and collapse spaces', int(untrimmed))

    blank_cities = int(customers['city'].isna().sum())
    customers['city'] = customers['city'].fillna('Unknown')
    note('Customers', 'city', 'Blank city',
         "Fill with 'Unknown' rather than dropping the customer — their orders still count",
         blank_cities)

    for col in ['customer_name', 'province', 'acquisition_channel']:
        customers[col] = _trim(customers[col])

    # --- Products -------------------------------------------------------
    products['category'] = _trim(products['category']).replace(CATEGORY_MAP)
    products['product_name'] = _trim(products['product_name'])
    note('Products', 'category', "'Ready-to-Drink' and 'Ready to Drink' split one category",
         "Map to 'Ready to Drink'",
         int(raw['Products']['category'].isin(CATEGORY_MAP.keys()).sum()))

    # --- Marketing_Spend ------------------------------------------------
    spend['channel'] = _trim(spend['channel'])

    note('Orders', 'order_status', "Only 'Completed' counts as revenue",
         "NOT removed — 'Returned' and 'Cancelled' rows are kept so you can measure "
         "the return rate. Filter them out when computing revenue or profit.", 0)

    return {'Orders': orders, 'Customers': customers, 'Products': products,
            'Marketing_Spend': spend}, pd.DataFrame(log)


def enrich(t):
    """Line-level table with the brief's revenue and profit formulas applied."""
    df = t['Orders'].merge(
        t['Products'][['product_id', 'product_name', 'category', 'unit_cost']],
        on='product_id', how='left')
    df = df.merge(
        t['Customers'][['customer_id', 'gender', 'age', 'city', 'acquisition_channel']]
        .rename(columns={'city': 'customer_city'}),  # Orders.city is the shipping city
        on='customer_id', how='left')

    net = df['unit_price'] * (1 - df['discount_pct'])
    df['net_unit_price'] = net.round(2)
    df['revenue'] = (df['quantity'] * net).round(2)
    df['cogs'] = (df['quantity'] * df['unit_cost']).round(2)
    df['profit'] = (df['revenue'] - df['cogs']).round(2)
    df['discount_idr'] = (df['quantity'] * df['unit_price'] * df['discount_pct']).round(2)
    df['is_completed'] = df['order_status'] == 'Completed'
    df['month'] = df['order_date'].dt.to_period('M').dt.to_timestamp()
    return df


# ── Sessions 3-5: the answers a mid-program joiner starts from ───────────────

def _margin(df):
    return (df['profit'] / df['revenue'] * 100).round(1)


def answers(enriched):
    """Finished Q1-Q4 and Q7 tables, completed orders only."""
    done = enriched[enriched['is_completed']]

    q1_cat = done.groupby('category', as_index=False).agg(
        orders=('order_id', 'nunique'), units=('quantity', 'sum'),
        revenue=('revenue', 'sum'), cogs=('cogs', 'sum'), profit=('profit', 'sum'))
    q1_cat['margin_pct'] = _margin(q1_cat)
    q1_cat = q1_cat.sort_values('profit', ascending=False)

    q1_prod = done.groupby(['product_id', 'product_name', 'category'], as_index=False).agg(
        units=('quantity', 'sum'), revenue=('revenue', 'sum'), profit=('profit', 'sum'))
    q1_prod['margin_pct'] = _margin(q1_prod)
    q1_prod = q1_prod.sort_values('profit', ascending=False)

    q2 = done.groupby('channel', as_index=False).agg(
        orders=('order_id', 'nunique'), customers=('customer_id', 'nunique'),
        revenue=('revenue', 'sum'), discount_given=('discount_idr', 'sum'),
        profit=('profit', 'sum'))
    q2['avg_order_value'] = (q2['revenue'] / q2['orders']).round(0)
    q2['margin_pct'] = _margin(q2)
    q2 = q2.sort_values('profit', ascending=False)

    q3 = done.groupby('month', as_index=False).agg(
        orders=('order_id', 'nunique'), customers=('customer_id', 'nunique'),
        units=('quantity', 'sum'), revenue=('revenue', 'sum'), profit=('profit', 'sum'))
    q3['avg_order_value'] = (q3['revenue'] / q3['orders']).round(0)
    q3 = q3.sort_values('month')

    per_customer = done.groupby('customer_id', as_index=False).agg(
        orders=('order_id', 'nunique'), revenue=('revenue', 'sum'))
    per_customer['buyer_type'] = per_customer['orders'].gt(1).map(
        {True: 'Repeat buyer', False: 'One-time buyer'})
    q4 = per_customer.groupby('buyer_type', as_index=False).agg(
        customers=('customer_id', 'count'), orders=('orders', 'sum'),
        revenue=('revenue', 'sum'))
    q4['customer_share_pct'] = (q4['customers'] / q4['customers'].sum() * 100).round(1)
    q4['revenue_share_pct'] = (q4['revenue'] / q4['revenue'].sum() * 100).round(1)

    q7 = done.groupby('discount_pct', as_index=False).agg(
        order_lines=('order_id', 'count'), avg_quantity=('quantity', 'mean'),
        units=('quantity', 'sum'), revenue=('revenue', 'sum'), profit=('profit', 'sum'))
    q7['avg_quantity'] = q7['avg_quantity'].round(2)
    q7['margin_pct'] = _margin(q7)

    return {
        'Q1_Category_Profit': q1_cat, 'Q1_Product_Profit': q1_prod,
        'Q2_Channel_Value': q2, 'Q3_Monthly_Trend': q3,
        'Q4_Repeat_Buyers': q4, 'Q7_Discount_vs_Volume': q7,
    }


# ── Session 10: RFM ──────────────────────────────────────────────────────────

def _segment(r, f):
    """Champions / Loyal / At-Risk / Lost, the four the brief names."""
    if r >= 4 and f >= 4:
        return 'Champions'
    if f >= 4:
        return 'Loyal'
    if r >= 4:
        return 'Potential Loyalist'
    if r <= 2 and f >= 3:
        return 'At-Risk'
    if r <= 2:
        return 'Lost'
    return 'Needs Attention'


def rfm(enriched, customers):
    done = enriched[enriched['is_completed']]
    agg = done.groupby('customer_id', as_index=False).agg(
        last_order_date=('order_date', 'max'),
        frequency=('order_id', 'nunique'),
        monetary=('revenue', 'sum'))
    agg['recency_days'] = (SNAPSHOT - agg['last_order_date']).dt.days
    agg['monetary'] = agg['monetary'].round(0)

    # Quintiles: recent, frequent and high-spending all score 5.
    agg['r_score'] = pd.qcut(agg['recency_days'], 5, labels=[5, 4, 3, 2, 1]).astype(int)
    agg['f_score'] = pd.qcut(agg['frequency'].rank(method='first'), 5, labels=[1, 2, 3, 4, 5]).astype(int)
    agg['m_score'] = pd.qcut(agg['monetary'], 5, labels=[1, 2, 3, 4, 5]).astype(int)
    agg['rfm_score'] = (agg['r_score'].astype(str) + agg['f_score'].astype(str)
                        + agg['m_score'].astype(str))
    agg['segment'] = [_segment(r, f) for r, f in zip(agg['r_score'], agg['f_score'])]

    out = customers[['customer_id', 'customer_name', 'city', 'acquisition_channel']].merge(
        agg, on='customer_id', how='left')

    # A customer with no completed order has no R/F/M to score. Kept in the file
    # rather than dropped so the row counts still reconcile against Customers —
    # filter them out before you size the segments.
    never = out['segment'].isna()
    out.loc[never, ['frequency', 'monetary']] = 0
    out.loc[never, 'segment'] = 'Never Purchased'

    out['last_order_date'] = out['last_order_date'].dt.strftime('%Y-%m-%d')
    # The left join widened the integer columns to float; Int64 keeps the blanks
    # blank without printing 112.0 where the learner expects 112.
    for col in ['recency_days', 'frequency', 'monetary', 'r_score', 'f_score', 'm_score']:
        out[col] = out[col].astype('Int64')
    out['rfm_score'] = (out['r_score'].astype('string') + out['f_score'].astype('string')
                        + out['m_score'].astype('string'))

    return out[['customer_id', 'customer_name', 'city', 'acquisition_channel',
                'last_order_date', 'recency_days', 'frequency', 'monetary',
                'r_score', 'f_score', 'm_score', 'rfm_score', 'segment']]


# ── The workbook's Project Brief tab ─────────────────────────────────────────

def patch_brief(path):
    """Mark S7 / S8 / S10 optional so the workbook tells the same story as the app."""
    wb = openpyxl.load_workbook(path)
    ws = wb['Project Brief']

    ws['C26'] = ('Q5.  Using RFM analysis, what customer segments exist (e.g. Champions, '
                 'Loyal, At-Risk, Lost)? How big and how valuable is each?   '
                 '[OPTIONAL — answered in Session 10]')

    # Q6 used to live only in Sessions 7 and 10. Both are optional now, so the
    # required path picks it up in Session 4, where the tables are already joined.
    ws['D38'] = ('Load the four tables into a database and write SQL (SELECT, WHERE, ORDER BY, '
                 'GROUP BY, HAVING, JOIN) to build the exact analysis datasets you need '
                 '— e.g. join Orders to Products to compute profit, or Orders to Customers '
                 'for segment analysis. Also join Customers.acquisition_channel to '
                 'Marketing_Spend to compare CAC and ROI per channel (answers Q6). Note that '
                 "the two channel vocabularies do not match 1:1 — 'Marketplace' vs "
                 "'Marketplace Ads', and Organic Social and Referral have no spend rows — so "
                 'state your mapping assumption.')

    ws['C32'] = ('Your final submission should demonstrate the following. Treat each as a '
                 'checklist for your portfolio. Sessions 7, 8 and 10 are OPTIONAL — the project '
                 'is complete without them; completing them adds an interactive dashboard, a '
                 'reproducible notebook and a customer segmentation to your portfolio.')

    ws['C41'] = 'PowerBI Dashboard   [OPTIONAL]'
    ws['D41'] = ('Assemble an interactive PowerBI dashboard: revenue trend, channel mix, '
                 'category profit, RFM segments (if you completed Session 10), and marketing '
                 'ROI. Publish it.')

    ws['C42'] = 'Python / Pandas 101   [OPTIONAL]'
    ws['D42'] = ('Reproduce the cleaning and key aggregations in pandas; visualize with '
                 'Matplotlib/Seaborn; optionally automate an Excel report with openpyxl.')

    ws['C44'] = f'Advanced Analysis {EM} Segmentation & RFM   [OPTIONAL]'
    ws['D44'] = ('Compute Recency, Frequency, Monetary per customer (snapshot date = 1 Jan 2026), '
                 'score and segment customers, and recommend a strategy per segment (answers Q5).')

    ws['C50'] = (f'{BULLET}   The data is RAW and intentionally messy {EM} cleaning it (Session 2) '
                 'is part of the assignment. Do NOT assume it is analysis-ready. If you joined the '
                 'program part-way through, each session page also offers a ready-to-continue file '
                 'so you can start from the correct state.')

    ws['C55'] = (f'{BULLET}   Deliver: (1) cleaned dataset, (2) SQL scripts, (3) pivot analysis, '
                 '(4) PowerBI dashboard link [optional], (5) Python notebook [optional], '
                 '(6) stakeholder deck, (7) 1-page executive summary answering Q8. The RFM '
                 'segmentation from Session 10 is optional too.')

    wb.save(path)


# ── Indonesian narrative variant ─────────────────────────────────────────────
#
# Learners who set the site to Bahasa Indonesia get an -id copy of each workbook
# that HAS narrative tabs (raw, cleaned, analysis). Only the two prose tabs are
# translated — 'Project Brief' and 'Data Dictionary'. Every data tab (Orders,
# Customers, Products, Marketing_Spend, the Q*/Cleaning_Log answer tabs) keeps
# its English column names untouched, because the SQL/pandas/Power BI exercises
# and the playgrounds all query those exact English identifiers.
#
# Coordinate-based, like patch_brief: the three source workbooks share one Brief
# and one Dictionary layout (cleaned/analysis are copied from RAW), so a single
# map localizes all three. Kept in lockstep with the English cells above and in
# the raw workbook — regenerate if either changes.

BRIEF_ID = {
    'C1': 'SEDUH COFFEE  —  PROYEK AKHIR DATA ANALYST',
    'C2': 'Proyek Portofolio Analisis Data End-to-End  —  Program Data Analyst Talentiv',
    'C5': f'1.  LATAR BELAKANG {EM} Tentang Seduh',
    'C6': ('Seduh adalah brand kopi spesialti direct-to-consumer (D2C) asal Indonesia yang tumbuh '
           'pesat, berdiri pada 2022. Perusahaan me-roasting dan menjual biji single-origin premium, '
           'kopi ready-to-drink (RTD), alat seduh, aksesori, langganan bulanan, dan gift set. Seduh '
           'berjualan online lewat webstore sendiri dan lewat marketplace (Tokopedia, Shopee, dan '
           'TikTok Shop), mengirim ke pelanggan di seluruh Indonesia.'),
    'C7': ('Pelanggan Seduh mayoritas profesional muda urban dan penggemar home-brewing berusia '
           '19-45 tahun. Brand ini tumbuh lewat iklan Instagram dan TikTok, referral, dan konten '
           'organik di media sosial.'),
    'C9': '2.  SITUASI BISNIS',
    'C10': ('Pada 2025, Seduh menghasilkan sekitar IDR 4,4 miliar revenue dari order yang selesai, '
            'naik dari sekitar IDR 2,0 miliar pada 2024. Di atas kertas bisnis ini bertumbuh '
            f'{EM} tetapi leadership khawatir:'),
    'C11': (f'{BULLET}   Marketing spend naik lebih cepat daripada revenue, sehingga biaya '
            'mengakuisisi tiap pelanggan terus meningkat.'),
    'C12': (f'{BULLET}   Banyak pelanggan hanya membeli sekali dan tidak pernah kembali. Repeat '
            f'purchase {EM} jantung dari bisnis kopi {EM} terasa lemah.'),
    'C13': (f'{BULLET}   Tidak ada yang tahu produk, channel, bulan, atau segmen pelanggan mana '
            'yang benar-benar mendorong profit (bukan sekadar revenue).'),
    'C14': 'Investor telah menetapkan target yang jelas untuk tahun depan:',
    'C15': ('    TARGET  ->   Tumbuh ke revenue IDR 5,5 miliar pada 2026 (sekitar +24% vs 2025), '
            'sambil meningkatkan retensi pelanggan dan efisiensi marketing.'),
    'C16': (f'Leadership merekrutmu {EM} sang data analyst {EM} untuk mengubah data operasional '
            'mentah Seduh menjadi diagnosis yang jelas dan rencana pertumbuhan yang bisa dijalankan.'),
    'C18': '3.  MISIMU',
    'C19': ('Dengan dataset yang tersedia di workbook ini (Orders, Customers, Products, '
            'Marketing_Spend), lakukan analisis end-to-end yang lengkap dan sampaikan rekomendasi '
            'berbasis data tentang bagaimana Seduh bisa mencapai target 2026-nya. Kamu akan '
            'menerapkan setiap skill dari 12 sesi program, mulai dari data cleaning hingga storytelling.'),
    'C21': '4.  PERTANYAAN BISNIS UTAMA YANG HARUS DIJAWAB',
    'C22': ('Q1.  Kategori produk dan produk individual mana yang paling PROFITABLE (revenue '
            'dikurangi biaya), bukan sekadar paling laku?'),
    'C23': ('Q2.  Channel penjualan mana (Webstore / Tokopedia / Shopee / TikTok Shop) yang '
            'paling bernilai setelah diskon?'),
    'C24': ('Q3.  Bagaimana tren revenue dari bulan ke bulan sepanjang 2024-2025? Apakah ada '
            'pola musiman (mis. Ramadan, Harbolnas 11.11 / 12.12)?'),
    'C25': ('Q4.  Berapa proporsi pelanggan yang membeli sekali vs repeat buyer? Berapa '
            'repeat-purchase rate-nya?'),
    'C26': ('Q5.  Dengan analisis RFM, segmen pelanggan apa saja yang ada (mis. Champions, Loyal, '
            f'At-Risk, Lost)? Seberapa besar dan seberapa bernilai masing-masing?   [OPSIONAL {EM} '
            'dijawab di Sesi 10]'),
    'C27': ('Q6.  Channel akuisisi mana yang mendatangkan pelanggan bernilai tertinggi? Apakah '
            'marketing spend sudah efisien (ROI / CAC)?'),
    'C28': (f'Q7.  Adakah hubungan antara pemberian diskon dan jumlah unit terjual {EM} dan apakah '
            'diskon benar-benar menaikkan profit?'),
    'C29': ('Q8.  Berdasarkan semua di atas, aksi spesifik apa yang harus diambil Seduh untuk '
            'mencapai target revenue IDR 5,5 M pada 2026?'),
    'C31': f'5.  DELIVERABLE {EM} Dipetakan ke 12 Sesi Program',
    'C32': ('Submission akhirmu harus menunjukkan hal-hal berikut. Perlakukan tiap poin sebagai '
            f'checklist untuk portofoliomu. Sesi 7, 8, dan 10 bersifat OPSIONAL {EM} proyek tetap '
            'lengkap tanpanya; menyelesaikannya menambah dashboard interaktif, notebook yang bisa '
            'diulang, dan segmentasi pelanggan ke portofoliomu.'),
    'B34': 'Sesi', 'C34': 'Deliverable', 'D34': 'Yang harus dihasilkan',
    'C35': 'Business Acumen & Problem Statement',
    'D35': ('Tulis problem statement yang jelas untuk Seduh dan petakan proses bisnisnya dari hulu '
            'ke hilir. Definisikan metrik yang penting di tiap tahap (akuisisi, konversi, order '
            'value, retensi). Tambahkan statistik deskriptif dasar dari data.'),
    'C36': 'Membersihkan & Merapikan Data',
    'D36': ('Perbaiki masalah kualitas pada data mentah: hapus order duplikat, tangani sel kosong, '
            'seragamkan teks yang tidak konsisten (nama channel, gender, city) memakai TRIM / '
            'CONCAT / LEFT / RIGHT, dan atasi nilai tidak valid (mis. quantity nol/negatif).'),
    'C37': 'Pivot Table untuk Insight',
    'D37': ('Gunakan Pivot Table untuk menjawab Q1-Q4: revenue & profit per kategori, per channel, '
            'per bulan, dan rincian repeat buyer.'),
    'C38': 'Pengumpulan Data dengan SQL',
    'D38': ('Muat keempat tabel ke database dan tulis SQL (SELECT, WHERE, ORDER BY, GROUP BY, '
            'HAVING, JOIN) untuk membangun dataset analisis persis seperti yang kamu butuhkan '
            f'{EM} mis. join Orders ke Products untuk menghitung profit, atau Orders ke Customers '
            'untuk analisis segmen. Join juga Customers.acquisition_channel ke Marketing_Spend '
            'untuk membandingkan CAC dan ROI per channel (menjawab Q6). Perhatikan bahwa kedua '
            f"kosakata channel tidak sama persis 1:1 {EM} 'Marketplace' vs 'Marketplace Ads', serta "
            f'Organic Social dan Referral tidak punya baris spend {EM} jadi nyatakan asumsi pemetaanmu.'),
    'C39': 'Exploratory Data Analysis (EDA)',
    'D39': ('Lakukan analisis univariat dan bivariat: sebaran order value, quantity, rating; '
            'hubungan antara diskon dan quantity (Q7), harga dan permintaan, rating dan repeat purchase.'),
    'C40': 'Visualisasi Data',
    'D40': ('Bangun grafik yang jelas dan didesain baik (jenis chart, warna, tata letak yang tepat) '
            'yang menceritakan tren revenue dan cerita segmen.'),
    'C41': 'Dashboard PowerBI   [OPSIONAL]',
    'D41': ('Rakit dashboard PowerBI interaktif: tren revenue, bauran channel, profit kategori, '
            'segmen RFM (jika kamu menyelesaikan Sesi 10), dan ROI marketing. Publikasikan.'),
    'C42': 'Python / Pandas 101   [OPSIONAL]',
    'D42': ('Reproduksi cleaning dan agregasi utama di pandas; visualkan dengan Matplotlib/Seaborn; '
            'opsional, otomasi laporan Excel dengan openpyxl.'),
    'C43': 'Analisis End-to-End dengan AI',
    'D43': ('Gunakan alur kerja AI/agentik untuk mempercepat sebagian analisis (mis. menghasilkan '
            'SQL, merangkum temuan, atau mengotomasi laporan yang berulang) dan dokumentasikan '
            'cara pemakaiannya.'),
    'C44': f'Advanced Analysis {EM} Segmentasi & RFM   [OPSIONAL]',
    'D44': ('Hitung Recency, Frequency, Monetary per pelanggan (tanggal snapshot = 1 Jan 2026), '
            'beri skor dan segmentasi pelanggan, dan rekomendasikan strategi per segmen (menjawab Q5).'),
    'C45': 'Storytelling dengan Data',
    'D45': ('Ubah analisis menjadi narasi yang siap untuk stakeholder: dari data -> insight -> '
            'rekomendasi, disusun sebagai deck untuk leadership Seduh.'),
    'C46': 'Proyek Akhir & Portofolio',
    'D46': ('Kemas semuanya menjadi karya portofolio profesional: data yang sudah bersih, skrip '
            'SQL, dashboard, deck, dan ringkasan eksekutif tertulis yang menjawab Q8.'),
    'C48': '6.  CATATAN DATA & SUBMISSION',
    'C49': (f'{BULLET}   Workbook ini berisi 4 tab data: Orders (level transaksi), Customers, '
            "Products, dan Marketing_Spend. Lihat tab 'Data Dictionary' untuk definisi setiap kolom."),
    'C50': (f'{BULLET}   Data ini MENTAH dan sengaja dibuat berantakan {EM} membersihkannya (Sesi 2) '
            'adalah bagian dari tugas. JANGAN anggap data sudah siap dianalisis. Jika kamu bergabung '
            'di tengah program, tiap halaman sesi juga menyediakan file siap-lanjut agar kamu bisa '
            'mulai dari kondisi yang benar.'),
    'C51': (f'{BULLET}   Revenue per baris order = quantity x unit_price x (1 - discount_pct). '
            'Profit per baris = quantity x (unit_price x (1 - discount_pct) - unit_cost). unit_cost '
            'ada di tab Products.'),
    'C52': (f"{BULLET}   Hanya order dengan status = 'Completed' yang dihitung sebagai revenue. "
            "Order 'Returned' dan 'Cancelled' harus dikecualikan dari revenue/profit (tetapi berguna "
            'untuk metrik return-rate).'),
    'C53': (f'{BULLET}   Tanggal snapshot / analisis RFM = 1 Januari 2026 (anggap semua data lengkap '
            'sampai 31 Desember 2025).'),
    'C54': (f'{BULLET}   Keys: Orders.customer_id -> Customers.customer_id ; Orders.product_id -> '
            'Products.product_id ; Marketing_Spend di-join lewat month + channel.'),
    'C55': (f'{BULLET}   Serahkan: (1) dataset bersih, (2) skrip SQL, (3) analisis pivot, (4) tautan '
            'dashboard PowerBI [opsional], (5) notebook Python [opsional], (6) deck stakeholder, '
            '(7) ringkasan eksekutif 1 halaman yang menjawab Q8. Segmentasi RFM dari Sesi 10 juga opsional.'),
    'C57': f'Semoga berhasil {EM} buat leadership Seduh melihat ceritanya di dalam data.',
}

# Column names (order_id, quantity, …) and table names (Orders, …) are real
# identifiers the exercises query, so they stay English; only the header labels,
# the Type words and the Description prose are translated.
DICT_ID = {
    'B1': 'KAMUS DATA',
    'B2': 'Tabel', 'C2': 'Kolom', 'D2': 'Tipe', 'E2': 'Deskripsi',
    'E3': 'ID unik dari baris order (catatan: data mentah memuat beberapa baris duplikat yang harus dihapus).',
    'E4': 'Tanggal order dibuat (2024-01-01 sampai 2025-12-31).',
    'E5': 'Foreign key ke Customers.customer_id.',
    'E6': 'Foreign key ke Products.product_id.',
    'E7': 'Jumlah unit yang dipesan. Data mentah memuat beberapa nilai 0/negatif tidak valid yang harus dibersihkan.',
    'E8': 'Harga jual per unit saat order, sebelum diskon.',
    'E9': 'Diskon yang diterapkan pada baris sebagai pecahan (0,10 = 10%).',
    'E10': 'Channel penjualan: Webstore, Tokopedia, Shopee, TikTok Shop. Nilai mentah punya kapitalisasi/ejaan tidak konsisten.',
    'E11': 'Metode pembayaran yang dipakai (GoPay, OVO, DANA, Bank Transfer, Credit Card, COD, ShopeePay, Paylater).',
    'E12': 'Kota pengiriman.',
    'E13': 'Provinsi pengiriman.',
    'E14': "Completed, Returned, atau Cancelled. Hanya 'Completed' yang dihitung sebagai revenue.",
    'E15': 'Rating dari pelanggan; kosong jika tidak diberi rating atau order belum selesai.',
    'E16': 'Primary key.',
    'E17': 'Nama lengkap pelanggan.',
    'E18': f'Nilai mentah tidak konsisten (Female/F/female/Perempuan, dll.) {EM} seragamkan.',
    'E19': 'Usia dalam tahun.',
    'E20': 'Kota domisili. Ada beberapa sel kosong / kapitalisasi tidak konsisten / spasi di belakang yang harus dibersihkan.',
    'E21': 'Provinsi domisili.',
    'E22': 'Tanggal pelanggan pertama kali mendaftar.',
    'E23': 'Bagaimana pelanggan pertama kali diakuisisi (Instagram Ads, TikTok Ads, Google Search, Referral, Organic Social, Marketplace).',
    'E24': 'Primary key.',
    'E25': 'Nama produk / deskripsi SKU.',
    'E26': 'Roasted Beans, Ready-to-Drink, Brewing Equipment, Accessories, Subscription, Gift Set (data mentah punya inkonsistensi penamaan yang harus diperbaiki).',
    'E27': 'Harga list standar per unit.',
    'E28': 'Harga pokok per unit (COGS). Dipakai untuk analisis profit/margin.',
    'E29': 'Tanggal produk pertama kali didaftarkan.',
    'E30': 'Hari pertama bulan untuk catatan spend.',
    'E31': 'Channel marketing (Instagram Ads, TikTok Ads, Google Search, Marketplace Ads).',
    'E32': 'Belanja iklan pada bulan itu di channel tersebut.',
    'E33': 'Jumlah impresi iklan yang tayang.',
    'E34': 'Jumlah klik iklan yang diterima.',
}

# The Type column repeats a small vocabulary; translate every cell that holds one.
TYPE_ID = {
    'Text': 'Teks', 'Date': 'Tanggal', 'Number': 'Angka',
    'Number (IDR)': 'Angka (IDR)', 'Number (0-1)': 'Angka (0-1)',
    'Number (1-5)': 'Angka (1-5)',
}


def write_id_variant(en_path):
    """Copy an English workbook to its -id sibling and localize the two prose tabs."""
    dest = en_path.with_name(en_path.stem + '-id' + en_path.suffix)
    shutil.copy(en_path, dest)
    wb = openpyxl.load_workbook(dest)

    brief = wb['Project Brief']
    for coord, text in BRIEF_ID.items():
        brief[coord] = text

    dictionary = wb['Data Dictionary']
    for coord, text in DICT_ID.items():
        dictionary[coord] = text
    # Type column (D3:D34) — map each English type word to its Indonesian label.
    for row in range(3, dictionary.max_row + 1):
        cell = dictionary[f'D{row}']
        if cell.value in TYPE_ID:
            cell.value = TYPE_ID[cell.value]

    wb.save(dest)
    return dest


# ── Writers ──────────────────────────────────────────────────────────────────

def write_workbook(dest, sheets, order):
    """Copy the raw workbook (for its Brief and Dictionary tabs) and replace the data."""
    shutil.copy(RAW, dest)
    with pd.ExcelWriter(dest, engine='openpyxl', mode='a',
                        if_sheet_exists='replace', datetime_format='yyyy-mm-dd') as xl:
        for name, df in sheets.items():
            df.to_excel(xl, sheet_name=name, index=False)

    # to_excel appends replaced sheets at the end; restore a readable order with
    # the narrative tabs (Brief, Dictionary) still leading.
    wb = openpyxl.load_workbook(dest)
    lead = [n for n in wb.sheetnames if n not in order]
    wb._sheets = [wb[n] for n in lead + order]
    for name in order:
        wb[name].freeze_panes = 'A2'
    wb.save(dest)


SCHEMA = """\
-- Seduh Coffee — schema for the cleaned dataset.
-- Load the CSVs in this order; the foreign keys depend on it.
--
--   sqlite3 seduh.db < schema.sql
--   sqlite3 seduh.db ".mode csv" ".import --skip 1 products.csv products" ...
--
-- Dates are stored as ISO text (YYYY-MM-DD) so SQLite date functions work.

DROP TABLE IF EXISTS orders;
DROP TABLE IF EXISTS marketing_spend;
DROP TABLE IF EXISTS customers;
DROP TABLE IF EXISTS products;

CREATE TABLE products (
  product_id  TEXT PRIMARY KEY,
  product_name TEXT NOT NULL,
  category    TEXT NOT NULL,
  base_price  INTEGER NOT NULL,
  unit_cost   INTEGER NOT NULL,
  launch_date TEXT NOT NULL
);

CREATE TABLE customers (
  customer_id         TEXT PRIMARY KEY,
  customer_name       TEXT NOT NULL,
  gender              TEXT NOT NULL,
  age                 INTEGER NOT NULL,
  city                TEXT NOT NULL,
  province            TEXT,
  signup_date         TEXT NOT NULL,
  acquisition_channel TEXT NOT NULL
);

CREATE TABLE orders (
  order_id       TEXT PRIMARY KEY,
  order_date     TEXT NOT NULL,
  customer_id    TEXT NOT NULL REFERENCES customers(customer_id),
  product_id     TEXT NOT NULL REFERENCES products(product_id),
  quantity       INTEGER NOT NULL,
  unit_price     INTEGER NOT NULL,
  discount_pct   REAL NOT NULL,
  channel        TEXT NOT NULL,
  payment_method TEXT NOT NULL,
  city           TEXT NOT NULL,
  province       TEXT NOT NULL,
  order_status   TEXT NOT NULL,
  rating         INTEGER          -- NULL means the order was never rated
);

CREATE TABLE marketing_spend (
  month       TEXT NOT NULL,
  channel     TEXT NOT NULL,
  spend_idr   INTEGER NOT NULL,
  impressions INTEGER NOT NULL,
  clicks      INTEGER NOT NULL,
  PRIMARY KEY (month, channel)
);

CREATE INDEX idx_orders_customer ON orders(customer_id);
CREATE INDEX idx_orders_product  ON orders(product_id);
CREATE INDEX idx_orders_date     ON orders(order_date);
CREATE INDEX idx_orders_status   ON orders(order_status);
"""

PACK_README = """\
Seduh Coffee — cleaned data pack
================================

This is the dataset AFTER the Session 2 cleaning, in the formats the later
sessions import from. Use it if you joined the program part-way through.

  products.csv, customers.csv, orders.csv, marketing_spend.csv
      The four cleaned tables. UTF-8, comma-separated, ISO dates.

  orders_enriched.csv
      orders joined to products and customers, with the brief's formulas
      already applied: net_unit_price, revenue, cogs, profit, discount_idr,
      is_completed, month. Convenient for Power BI — but build it yourself
      at least once so you know what is in it.

  schema.sql
      Table definitions with keys and indexes.

  seduh.db
      SQLite database with all four tables already loaded. Open it with
      DB Browser for SQLite, or:  sqlite3 seduh.db

WHAT WAS CLEANED
  See the Cleaning_Log tab in seduh-coffee-cleaned.xlsx — every rule, why it
  was applied, and how many rows it touched.

WHAT WAS NOT
  - Blank ratings are still blank. An unrated order is not an error.
  - Returned and Cancelled orders are still here. Only 'Completed' counts as
    revenue, but you need the others to measure the return rate.

REMEMBER
  Revenue per line = quantity x unit_price x (1 - discount_pct)
  Profit  per line = revenue - (quantity x unit_cost)
  RFM snapshot date = 1 January 2026
"""


def _csv_text(df):
    """LF-terminated CSV. Anything else leaves a stray \\r on the last column,
    which silently turned every blank orders.rating into 0 in the playground."""
    return df.to_csv(index=False, date_format='%Y-%m-%d', lineterminator='\n')


def write_data_pack(dest, tables, enriched):
    csvs = {
        'products.csv': tables['Products'],
        'customers.csv': tables['Customers'],
        'orders.csv': tables['Orders'],
        'marketing_spend.csv': tables['Marketing_Spend'],
        'orders_enriched.csv': enriched,
    }

    # The four cleaned tables are also served standalone under public/project/data/
    # so both playgrounds can load the real dataset — a learner's playground totals
    # then match their own workbook exactly. The SQL playground builds its SQLite
    # database in the browser from these same CSVs, so there is one source of truth
    # and nothing ships twice.
    live = OUT / 'data'
    live.mkdir(exist_ok=True)

    tmp = OUT / '_seduh_tmp.db'
    tmp.unlink(missing_ok=True)
    con = sqlite3.connect(tmp)
    con.executescript(SCHEMA)
    for name, table in [('products', 'Products'), ('customers', 'Customers'),
                        ('orders', 'Orders'), ('marketing_spend', 'Marketing_Spend')]:
        df = tables[table].copy()
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%Y-%m-%d')
        df.to_sql(name, con, if_exists='append', index=False)
    con.commit()
    con.close()

    with zipfile.ZipFile(dest, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr('README.txt', PACK_README)
        z.writestr('schema.sql', SCHEMA)
        for name, df in csvs.items():
            z.writestr(name, _csv_text(df))
        z.write(tmp, 'seduh.db')

    tmp.unlink()

    # The playgrounds split these on ',' rather than carrying a CSV parser, which
    # is only safe while no value contains a comma or a quote. Fail loudly here
    # instead of corrupting a table in the browser.
    for name in ['products.csv', 'customers.csv', 'orders.csv', 'marketing_spend.csv']:
        text = _csv_text(csvs[name])
        if '"' in text or any(',' in str(v) for v in csvs[name].select_dtypes('object').to_numpy().ravel()):
            raise SystemExit(f'{name} contains a quoted or comma-bearing field; '
                             'the playground CSV loader cannot parse it')
        if '\r' in text:
            raise SystemExit(f'{name} contains a carriage return')
        # newline='' keeps Python from translating the LFs back into CRLFs.
        (live / name).write_text(text, encoding='utf-8', newline='')

    # orders_enriched stays zip-only so nobody skips building it themselves.


DECK_OUTLINE = """\
# Seduh Coffee — stakeholder deck outline

A skeleton for Session 11. One message per slide, and the message goes in the
title: a stakeholder who reads only the titles should still get the argument.
Every slide follows **data → insight → recommendation**.

Numbers below are placeholders — fill them from your own analysis. If you
disagree with a framing, change it. This is a starting point, not an answer key.

---

## 1. Title
Seduh Coffee — from IDR 4.4B to IDR 5.5B: where the growth comes from.
Your name, the date, and the one-line verdict.

## 2. The ask
Leadership wants +24% revenue in 2026 while improving retention and marketing
efficiency. State the three worries from the brief in one line each.

## 3. Where Seduh stands today
2024 → 2025 revenue, the growth rate, and the two things underneath it that are
moving the wrong way: rising acquisition cost and weak repeat purchase.

## 4. How I got here *(method, kept short)*
Four tables, ~24k order lines, cleaned per a documented log. State your
assumptions on one line: completed orders only, 1 Jan 2026 snapshot.

## 5. Q1 — Profit is not where volume is
Category and product profit vs revenue. Name the category that sells well and
earns badly, and the one that quietly carries margin.

## 6. Q2 — Not all channels are worth the same
Channel comparison **after** discount. If a channel looks big on gross revenue
and small on profit, that is the slide's message.

## 7. Q3 — The shape of the year
Monthly trend across 2024–2025 with Ramadan and Harbolnas 11.11 / 12.12
annotated. Say what is seasonal and what is real growth.

## 8. Q7 — Does discounting actually work?
Discount depth vs average quantity vs margin. Answer plainly: at what depth
does discounting stop paying for itself?

## 9. Q4 — The retention problem, sized
One-time vs repeat buyers: share of customers, share of revenue. This is the
slide the rest of the recommendation hangs on.

## 10. Q6 — Where the marketing money goes
Spend, CAC and ROI by acquisition channel. State your channel-mapping
assumption on the slide — the spend and customer vocabularies do not match 1:1.

## 11. *(Optional — Session 10)* Q5 — Who the customers actually are
RFM segments: Champions, Loyal, At-Risk, Lost. Size and value each. Skip this
slide if you did not do Session 10; the argument still holds without it.

## 12. Q8 — The plan
Three to five actions, each traced to a slide above. For each: what to do, what
it is worth in IDR, and how you would measure it.

## 13. The arithmetic
Bridge from IDR 4.4B to IDR 5.5B. Show each action's contribution adding up.
If it does not add up, say so — an honest gap beats a fabricated one.

## 14. What would have to be true
The assumptions your plan rests on, and what would falsify them. Anticipating
the pushback is what separates an analyst from a chart generator.

## 15. Appendix
Method notes, the cleaning log, and the charts you did not need in the main
line but will be asked about.

---

## One-page executive summary (Session 12)

Separate document, one page, no charts required:

1. **The situation** — two sentences.
2. **What the data says** — three findings, one line each, each with a number.
3. **What to do** — the three-to-five actions with their IDR contribution.
4. **What it adds up to** — the bridge to 5.5B.
5. **Assumptions and caveats** — completed-orders-only, the snapshot date, the
   channel mapping, and anything you had to decide during cleaning.
"""


def main():
    if not RAW.exists():
        raise SystemExit(f'missing {RAW}')

    patch_brief(RAW)
    print(f'patched  {RAW.name}  (Sessions 7/8/10 marked optional)')

    raw = pd.read_excel(RAW, sheet_name=DATA_SHEETS)
    tables, log = clean(raw)
    enriched = enrich(tables)
    answered = answers(enriched)
    print(f'cleaned  {len(raw["Orders"])} -> {len(tables["Orders"])} order rows, '
          f'{len(log)} logged rules')

    cleaned_path = OUT / 'seduh-coffee-cleaned.xlsx'
    write_workbook(cleaned_path,
                   {**tables, 'Cleaning_Log': log},
                   ['Cleaning_Log'] + DATA_SHEETS)

    analysis_path = OUT / 'seduh-coffee-analysis.xlsx'
    write_workbook(analysis_path,
                   {**tables, 'Cleaning_Log': log, 'Orders_Enriched': enriched, **answered},
                   ['Cleaning_Log'] + list(answered) + ['Orders_Enriched'] + DATA_SHEETS)

    pack_path = OUT / 'seduh-coffee-data-pack.zip'
    write_data_pack(pack_path, tables, enriched)

    rfm_path = OUT / 'seduh-coffee-rfm-segments.csv'
    rfm(enriched, tables['Customers']).to_csv(rfm_path, index=False)

    deck_path = OUT / 'seduh-coffee-deck-outline.md'
    deck_path.write_text(DECK_OUTLINE, encoding='utf-8')

    # Indonesian variants: only the workbooks with narrative tabs get one. The
    # data pack, RFM CSV and deck outline have no Project Brief / Data Dictionary
    # to translate, so they stay language-neutral.
    id_paths = [write_id_variant(p) for p in [RAW, cleaned_path, analysis_path]]
    print(f'localized {len(id_paths)} workbook(s) to Bahasa Indonesia (Brief + Dictionary only)')

    print()
    for p in [RAW, cleaned_path, pack_path, analysis_path, rfm_path, deck_path,
              *id_paths, *sorted((OUT / 'data').iterdir())]:
        print(f'  {p.stat().st_size / 1_048_576:6.2f} MB  {p.relative_to(OUT)}')


if __name__ == '__main__':
    main()
